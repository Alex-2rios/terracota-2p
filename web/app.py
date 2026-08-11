"""Panel web de administración de Terracota.

Consume exclusivamente la API REST (`api_client.TerracotaApi`). No importa
psycopg ni conoce la cadena de conexión a PostgreSQL: si necesitas un dato
nuevo, se agrega un endpoint en la API, no una consulta aquí.
"""

from __future__ import annotations

import io
import os
import secrets
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps
from xml.sax.saxutils import escape as escapar_xml

import openpyxl
from flask import (
    Flask, flash, jsonify, redirect, render_template, request, send_file, session, url_for,
)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from api_client import ApiError, TerracotaApi

app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config["API_URL"] = os.environ.get("API_URL", "http://localhost:8080/api/v1")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

api = TerracotaApi(app.config["API_URL"])

RUTAS_PUBLICAS = {"login", "static", "salud"}

ESTADOS_PEDIDO = ["PENDIENTE", "PREPARANDO", "LISTO", "ENTREGADO", "PAGADO", "CANCELADO"]

ETIQUETA_ESTADO_INVENTARIO = {
    "DISPONIBLE": "Disponible",
    "BAJO": "Bajo stock",
    "AGOTADO": "Agotado",
    "NO_DISPONIBLE": "No disponible",
    "ELIMINADO": "Dado de baja",
}

def _cerrar_sesion(mensaje: str | None = None, categoria: str = "error"):
    session.clear()
    if mensaje:
        flash(mensaje, categoria)
    return redirect(url_for("login"))

@app.before_request
def exigir_sesion_y_csrf():
    if request.endpoint in RUTAS_PUBLICAS or request.endpoint is None:
        return None

    if not session.get("token"):
        return _cerrar_sesion("Inicia sesión para continuar.", "error")

    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        enviado = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
        if not enviado or not secrets.compare_digest(enviado, session.get("csrf_token", "")):
            flash("La sesión del formulario expiró. Vuelve a intentarlo.", "error")
            return redirect(request.referrer or url_for("inicio"))
    return None

@app.context_processor
def inyectar_layout():
    return {
        "admin_user": session.get("nombre", "Administrador"),
        "admin_email": session.get("usuario", ""),
        "csrf_token": session.get("csrf_token", ""),
        "anio_actual": date.today().year,
    }

def manejar_errores_api(vista):
    """Convierte cualquier ApiError en algo presentable, nunca en un 500.

    Se distingue el error del usuario (filtros mal puestos, recurso que no
    existe) de la caída del servicio: no es lo mismo "revisa las fechas" que
    "la API no responde", y mostrar lo segundo cuando pasa lo primero confunde.
    """

    @wraps(vista)
    def envoltura(*args, **kwargs):
        try:
            return vista(*args, **kwargs)
        except ApiError as error:
            if error.es_sesion_invalida:
                return _cerrar_sesion("Tu sesión expiró. Inicia sesión de nuevo.")

            codigo = error.status_code or 0
            es_culpa_del_usuario = 400 <= codigo < 500

            if request.method == "POST":
                flash(error.mensaje, "error")
                destino = request.referrer
                if destino and destino != request.url:
                    return redirect(destino)

            if es_culpa_del_usuario and request.query_string:
                flash(error.mensaje, "error")
                return redirect(request.path)

            return render_template(
                "error.html",
                page_title="No se pudo mostrar",
                active_page=None,
                titulo="No se pudo cargar la información" if es_culpa_del_usuario
                       else "El servicio no está disponible",
                mensaje=error.mensaje,
                detalle=None if es_culpa_del_usuario
                        else f"La API configurada es {app.config['API_URL']}.",
                mostrar_ayuda=not es_culpa_del_usuario,
                reintentar=url_for("inicio") if es_culpa_del_usuario else request.url,
            ), (codigo if es_culpa_del_usuario else 503)

    return envoltura

def token() -> str:
    return session["token"]

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("correo", "").strip()
        password = request.form.get("password", "")

        if not usuario or not password:
            flash("Ingresa tu usuario y contraseña.", "error")
            return redirect(url_for("login"))

        try:
            sesion = api.login(usuario, password)
        except ApiError as error:
            flash(error.mensaje, "error")
            return redirect(url_for("login"))

        identidad = sesion["usuario"]
        if "administrador" not in identidad["roles"]:
            flash("Acceso denegado: el panel web es exclusivo del rol Administrador.", "error")
            return redirect(url_for("login"))

        session.clear()
        session.permanent = True
        session["token"] = sesion["access_token"]
        session["user_id"] = identidad["id"]
        session["nombre"] = identidad["nombre"]
        session["usuario"] = identidad["usuario"]
        session["roles"] = identidad["roles"]
        session["csrf_token"] = secrets.token_urlsafe(32)
        return redirect(url_for("inicio"))

    mensajes = session.get("_flashes")
    session.clear()
    if mensajes:
        session["_flashes"] = mensajes

    estado_api = "ok"
    try:
        api.health()
    except ApiError as error:
        estado_api = error.mensaje

    return render_template(
        "login.html", page_title="Iniciar Sesión", hide_nav=True,
        api_url=app.config["API_URL"], estado_api=estado_api,
    )

@app.post("/salir")
def salir():
    return _cerrar_sesion("Sesión cerrada correctamente.", "success")

@app.get("/salud")
def salud():
    """Comprueba que la web puede hablar con la API (útil en la revisión)."""
    try:
        return jsonify({"web": "ok", "api": api.health()})
    except ApiError as error:
        return jsonify({"web": "ok", "api": "error", "detalle": error.mensaje}), 503

@app.get("/inicio")
@manejar_errores_api
def inicio():
    datos = api.dashboard(token())

    metricas = [
        {"icon": "ventas-mes.png", "label": "Ventas del Mes", "value": _dinero(datos["ventas_mes"])},
        {"icon": "pedidos-hoy.png", "label": "Pedidos Hoy", "value": str(datos["pedidos_hoy"])},
        {"icon": "usuario.png", "label": "Usuarios Activos", "value": str(datos["usuarios_activos"])},
        {"icon": "crecimiento.png", "label": "Ventas de Hoy", "value": _dinero(datos["ventas_hoy"])},
    ]

    return render_template(
        "inicio.html",
        page_title="Estadísticas",
        active_page="inicio",
        metrics=metricas,
        pedidos=datos["pedidos_recientes"],
        chart_sales=datos["serie_7_dias"],
        chart_products=datos["top_productos"],
        bajo_stock=datos["productos_bajo_stock"],
    )

@app.get("/estadisticas")
def estadisticas():
    return redirect(url_for("inicio"))

@app.get("/pedidos")
@manejar_errores_api
def pedidos():
    hoy = date.today()
    fecha_inicio = request.args.get("fecha_inicio") or (hoy - timedelta(days=7)).isoformat()
    fecha_fin = request.args.get("fecha_fin") or hoy.isoformat()
    estado_sel = request.args.get("estado", "Todos")
    mesa_sel = request.args.get("mesa", "")

    listado = api.pedidos(
        token(),
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        estado=None if estado_sel == "Todos" else estado_sel,
        mesa=_entero(mesa_sel),
    )

    return render_template(
        "pedidos.html",
        page_title="Pedidos",
        active_page="pedidos",
        pedidos=listado,
        estados=ESTADOS_PEDIDO,
        fecha_inicio_str=fecha_inicio,
        fecha_fin_str=fecha_fin,
        estado_sel=estado_sel,
        mesa_sel=mesa_sel,
    )

@app.get("/pedidos/<int:pedido_id>")
def detalle_pedido(pedido_id: int):
    """Devuelve JSON para el modal de detalle."""
    try:
        return jsonify(api.pedido(token(), pedido_id))
    except ApiError as error:
        return jsonify({"error": error.mensaje}), error.status_code or 500

@app.post("/pedidos/<int:pedido_id>/cancelar")
@manejar_errores_api
def cancelar_pedido(pedido_id: int):
    motivo = request.form.get("motivo", "").strip() or "Cancelado por el administrador desde el panel web."
    resultado = api.cancelar_pedido(token(), pedido_id, motivo)
    flash(resultado.get("mensaje", "Pedido cancelado."), "success")
    return redirect(url_for("pedidos", **request.args.to_dict()))

@app.get("/usuarios")
@manejar_errores_api
def usuarios():
    buscar = request.args.get("buscar", "").strip()
    incluir = request.args.get("incluir_eliminados") == "1"

    listado = api.usuarios(token(), buscar=buscar, incluir_eliminados="true" if incluir else None)
    for usuario in listado:
        usuario["estado"] = (
            "Baja" if usuario["eliminado"] else ("Activo" if usuario["activo"] else "Inactivo")
        )
        usuario["roles_texto"] = ", ".join(usuario["roles_nombre"]) or "Sin rol"
        usuario["ultimo_acceso_texto"] = _fecha_corta(usuario.get("ultimo_acceso")) or "Nunca"

    return render_template(
        "usuarios.html",
        page_title="Gestión de Usuarios",
        active_page="usuarios",
        usuarios=listado,
        buscar=buscar,
        incluir_eliminados=incluir,
    )

@app.route("/usuarios/nuevo", methods=["GET", "POST"])
@manejar_errores_api
def agregar_usuario():
    roles = api.roles(token())

    if request.method == "POST":
        error = _validar_password(request.form.get("password", ""), request.form.get("confirmacion", ""))
        if error:
            flash(error, "error")
            return redirect(url_for("agregar_usuario"))

        seleccionados = request.form.getlist("roles")
        if not seleccionados:
            flash("Selecciona al menos un rol.", "error")
            return redirect(url_for("agregar_usuario"))

        api.crear_usuario(token(), {
            "nombre": request.form.get("nombre", "").strip(),
            "usuario": request.form.get("usuario", "").strip(),
            "password": request.form.get("password", ""),
            "roles": seleccionados,
            "activo": request.form.get("activo", "true") == "true",
        })
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("usuarios"))

    return render_template(
        "usuario_form.html",
        page_title="Agregar Usuario",
        active_page="usuarios",
        action_label="Agregar Usuario",
        roles=roles,
        usuario={"id": 0, "nombre": "", "usuario": "", "roles": [], "activo": True, "eliminado": False},
        es_nuevo=True,
    )

@app.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@manejar_errores_api
def editar_usuario(user_id: int):
    roles = api.roles(token())

    if request.method == "POST":
        password = request.form.get("password", "")
        confirmacion = request.form.get("confirmacion", "")

        if password or confirmacion:
            error = _validar_password(password, confirmacion)
            if error:
                flash(error, "error")
                return redirect(url_for("editar_usuario", user_id=user_id))

        seleccionados = request.form.getlist("roles")
        if not seleccionados:
            flash("Selecciona al menos un rol.", "error")
            return redirect(url_for("editar_usuario", user_id=user_id))

        cambios = {
            "nombre": request.form.get("nombre", "").strip(),
            "usuario": request.form.get("usuario", "").strip(),
            "activo": request.form.get("activo", "true") == "true",
            "roles": seleccionados,
        }
        if password:
            cambios["password"] = password

        api.actualizar_usuario(token(), user_id, cambios)
        flash("Usuario actualizado correctamente.", "success")
        return redirect(url_for("usuarios"))

    return render_template(
        "usuario_form.html",
        page_title="Editar Usuario",
        active_page="usuarios",
        action_label="Guardar Cambios",
        roles=roles,
        usuario=api.usuario(token(), user_id),
        es_nuevo=False,
    )

@app.post("/usuarios/<int:user_id>/eliminar")
@manejar_errores_api
def eliminar_usuario(user_id: int):
    resultado = api.eliminar_usuario(token(), user_id)
    flash(resultado.get("mensaje", "Usuario dado de baja."), "success")
    return redirect(url_for("usuarios"))

@app.post("/usuarios/<int:user_id>/reactivar")
@manejar_errores_api
def reactivar_usuario(user_id: int):
    api.reactivar_usuario(token(), user_id)
    flash("Usuario reactivado.", "success")
    return redirect(url_for("usuarios", incluir_eliminados="1"))

@app.get("/inventario")
@manejar_errores_api
def inventario():
    buscar = request.args.get("buscar", "").strip()
    categoria = request.args.get("categoria", "Todos")
    incluir = request.args.get("incluir_eliminados") == "1"

    productos = api.inventario(
        token(), buscar=buscar, categoria=categoria,
        incluir_eliminados="true" if incluir else None,
    )
    for producto in productos:
        producto["estado_texto"] = ETIQUETA_ESTADO_INVENTARIO.get(producto["estado"], producto["estado"])
        producto["precio_texto"] = _dinero(producto["precio"])

    return render_template(
        "inventario.html",
        page_title="Gestión de Inventario",
        active_page="inventario",
        productos=productos,
        alertas=api.alertas_inventario(token()),
        categorias=api.categorias(token()),
        categoria_sel=categoria,
        buscar=buscar,
        incluir_eliminados=incluir,
    )

@app.route("/inventario/nuevo", methods=["GET", "POST"])
@manejar_errores_api
def agregar_producto():
    categorias = api.categorias(token())

    if request.method == "POST":
        datos, error = _leer_producto(request.form)
        if error:
            flash(error, "error")
            return redirect(url_for("agregar_producto"))

        api.crear_producto(token(), datos)
        flash("Producto creado correctamente.", "success")
        return redirect(url_for("inventario"))

    return render_template(
        "producto_form.html",
        page_title="Agregar Producto",
        active_page="inventario",
        action_label="Agregar Producto",
        categorias=categorias,
        producto={
            "id": 0, "nombre": "", "categoria": categorias[0]["nombre"] if categorias else "",
            "descripcion": "", "stock_actual": 50, "stock_minimo": 15, "precio": "", "disponible": True,
        },
        es_nuevo=True,
    )

@app.route("/inventario/<int:product_id>/editar", methods=["GET", "POST"])
@manejar_errores_api
def editar_producto(product_id: int):
    categorias = api.categorias(token())

    if request.method == "POST":
        datos, error = _leer_producto(request.form)
        if error:
            flash(error, "error")
            return redirect(url_for("editar_producto", product_id=product_id))

        api.actualizar_producto(token(), product_id, datos)
        flash("Producto actualizado correctamente.", "success")
        return redirect(url_for("inventario"))

    return render_template(
        "producto_form.html",
        page_title="Editar Producto",
        active_page="inventario",
        action_label="Guardar Cambios",
        categorias=categorias,
        producto=api.producto(token(), product_id),
        es_nuevo=False,
    )

@app.post("/inventario/<int:product_id>/eliminar")
@manejar_errores_api
def eliminar_producto(product_id: int):
    resultado = api.eliminar_producto(token(), product_id)
    flash(resultado.get("mensaje", "Producto dado de baja."), "success")
    return redirect(url_for("inventario"))

@app.route("/gastos", methods=["GET", "POST"])
@manejar_errores_api
def gastos():
    if request.method == "POST":
        monto = _decimal(request.form.get("monto", ""))
        if monto is None or monto < 0:
            flash("Captura un monto válido.", "error")
            return redirect(url_for("gastos"))

        api.crear_gasto(token(), {
            "concepto": request.form.get("concepto", "").strip(),
            "monto": str(monto),
            "fecha": request.form.get("fecha") or None,
        })
        flash("Gasto registrado.", "success")
        return redirect(url_for("gastos"))

    hoy = date.today()
    fecha_inicio = request.args.get("fecha_inicio") or hoy.replace(day=1).isoformat()
    fecha_fin = request.args.get("fecha_fin") or hoy.isoformat()

    listado = api.gastos(token(), fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    total = sum(Decimal(str(g["monto"])) for g in listado)

    return render_template(
        "gastos.html",
        page_title="Gastos",
        active_page="gastos",
        gastos=listado,
        total=_dinero(total),
        fecha_inicio_str=fecha_inicio,
        fecha_fin_str=fecha_fin,
        hoy=hoy.isoformat(),
    )

@app.post("/gastos/<int:gasto_id>/eliminar")
@manejar_errores_api
def eliminar_gasto(gasto_id: int):
    api.eliminar_gasto(token(), gasto_id)
    flash("Gasto eliminado.", "success")
    return redirect(url_for("gastos"))

@app.get("/reportes")
@manejar_errores_api
def reportes():
    hoy = date.today()
    return render_template(
        "reportes.html",
        page_title="Reportes",
        active_page="reportes",
        catalogo=api.opciones_reporte(token()),
        fecha_inicio=(hoy - timedelta(days=30)).isoformat(),
        fecha_fin=hoy.isoformat(),
    )

def _filtros_del_formulario() -> dict:
    """Traduce el formulario a los parámetros de la API.

    Los valores comodín ("Todos", "Todas") se omiten para que la API no filtre
    por ellos, y `mesa` sólo viaja si de verdad es un número.
    """
    comodines = {"", "todos", "todas"}
    crudo = {
        "fecha_inicio": request.form.get("fecha_inicio"),
        "fecha_fin": request.form.get("fecha_fin"),
        "categoria": request.form.get("categoria"),
        "estado": request.form.get("estado"),
        "usuario": request.form.get("usuario"),
        "metodo": request.form.get("metodo"),
    }
    filtros = {
        clave: valor for clave, valor in crudo.items()
        if valor and valor.strip().lower() not in comodines
    }
    mesa = _entero(request.form.get("mesa", ""))
    if mesa is not None:
        filtros["mesa"] = mesa
    return filtros

@app.post("/reportes/ver")
@manejar_errores_api
def ver_reporte():
    """Muestra el reporte en pantalla antes de descargarlo."""
    tipo = request.form.get("tipo_reporte", "ventas")
    reporte = api.reporte(token(), tipo=tipo, **_filtros_del_formulario())
    return render_template(
        "reporte_vista.html",
        page_title=reporte["titulo"],
        active_page="reportes",
        reporte=reporte,
        tipo=tipo,
        filtros=request.form.to_dict(),
    )

@app.post("/reportes/exportar")
@manejar_errores_api
def exportar_reporte():
    tipo = request.form.get("tipo_reporte", "ventas")
    formato = request.form.get("formato", "pdf").lower()

    reporte = api.reporte(token(), tipo=tipo, **_filtros_del_formulario())

    marca = datetime.now().strftime("%Y%m%d_%H%M")
    if formato == "xlsx":
        return send_file(
            _construir_xlsx(reporte), as_attachment=True,
            download_name=f"reporte_{tipo}_{marca}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return send_file(
        _construir_pdf(reporte), as_attachment=True,
        download_name=f"reporte_{tipo}_{marca}.pdf",
        mimetype="application/pdf",
    )

def _dinero(valor) -> str:
    try:
        return f"${Decimal(str(valor or 0)):,.2f}"
    except (InvalidOperation, TypeError):
        return "$0.00"

def _entero(valor: str) -> int | None:
    try:
        return int(str(valor).strip())
    except (TypeError, ValueError):
        return None

def _decimal(valor: str) -> Decimal | None:
    limpio = str(valor or "").replace("$", "").replace(",", "").strip()
    try:
        return Decimal(limpio)
    except (InvalidOperation, TypeError):
        return None

def _fecha_corta(valor: str | None) -> str | None:
    if not valor:
        return None
    try:
        return datetime.fromisoformat(str(valor)).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return str(valor)[:16]

def _validar_password(password: str, confirmacion: str) -> str | None:
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres."
    if password != confirmacion:
        return "La contraseña y su confirmación no coinciden."
    return None

def _leer_producto(form) -> tuple[dict, str | None]:
    """Convierte el formulario en el cuerpo que espera la API, sin reventar
    con campos vacíos o texto en los números."""
    nombre = form.get("nombre", "").strip()
    if len(nombre) < 2:
        return {}, "El nombre del producto es obligatorio."

    stock_actual = _entero(form.get("stock_actual", ""))
    stock_minimo = _entero(form.get("stock_minimo", ""))
    precio = _decimal(form.get("precio", ""))

    if stock_actual is None or stock_actual < 0:
        return {}, "El stock actual debe ser un número entero de cero o más."
    if stock_minimo is None or stock_minimo < 0:
        return {}, "El stock mínimo debe ser un número entero de cero o más."
    if precio is None or precio < 0:
        return {}, "El precio debe ser un número de cero o más."

    return {
        "nombre": nombre,
        "categoria": form.get("categoria", "").strip(),
        "descripcion": form.get("descripcion", "").strip() or None,
        "stock_actual": stock_actual,
        "stock_minimo": stock_minimo,
        "precio": str(precio),
        "disponible": form.get("disponible", "true") == "true",
    }, None

def _texto_pdf(valor) -> str:
    """Prepara un texto para reportlab.

    `Paragraph` interpreta un mini-HTML propio: sin escapar, un producto
    llamado «Café < 200 ml» se imprime como «Café» porque todo lo que va entre
    `<` y `>` se toma por una etiqueta y desaparece. Es pérdida de datos
    silenciosa dentro de un documento que se entrega.
    """
    return escapar_xml(str(valor))

INICIOS_DE_FORMULA = ("=", "+", "-", "@")

def _escribir_celda(celda, valor) -> None:
    """Escribe respetando el texto: un concepto como «=1+1» debe verse tal cual,
    no evaluarse como fórmula (además de ser el vector clásico de inyección).

    Al asignar `.value`, openpyxl marca por su cuenta como fórmula lo que
    empieza por «=»; se corrige el tipo justo después. (`set_explicit_value`
    ya no existe en openpyxl 3.1.)
    """
    celda.value = valor
    if isinstance(valor, str) and valor.startswith(INICIOS_DE_FORMULA):
        celda.data_type = "s"

def _construir_pdf(reporte: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=30, leftMargin=30, topMargin=34, bottomMargin=30,
        title=reporte["titulo"], author="Terracota",
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloTerracota", parent=estilos["Heading1"], fontSize=16,
        textColor=colors.HexColor("#6e321f"), spaceAfter=4,
    )
    estilo_sub = ParagraphStyle(
        "SubTerracota", parent=estilos["Normal"], fontSize=8,
        textColor=colors.HexColor("#8a7567"), spaceAfter=14,
    )
    estilo_seccion = ParagraphStyle(
        "SeccionTerracota", parent=estilos["Heading2"], fontSize=11,
        textColor=colors.HexColor("#73351f"), spaceBefore=12, spaceAfter=6,
    )
    estilo_celda = ParagraphStyle(
        "CeldaTerracota", parent=estilos["Normal"], fontSize=7.5, leading=9.5,
    )

    historia = [
        Paragraph(_texto_pdf(reporte["titulo"]), estilo_titulo),
        Paragraph(f"Generado el {_texto_pdf(reporte['generado_en'])} · Terracota cocina artesanal", estilo_sub),
    ]

    ancho_util = letter[0] - 60
    for seccion in reporte["secciones"]:
        historia.append(Paragraph(_texto_pdf(seccion["titulo"]), estilo_seccion))
        columnas = len(seccion["headers"])
        datos = [[Paragraph(f"<b>{_texto_pdf(h)}</b>", estilo_celda) for h in seccion["headers"]]]
        for fila in seccion["rows"]:
            datos.append([Paragraph(_texto_pdf(celda), estilo_celda) for celda in fila])

        tabla = Table(datos, colWidths=[ancho_util / columnas] * columnas, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#cdb7a5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2e211c")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFDF8")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFDF8"), colors.HexColor("#F7EFE6")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2d5c8")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        historia.append(tabla)
        historia.append(Spacer(1, 12))

    doc.build(historia)
    buffer.seek(0)
    return buffer

def _construir_xlsx(reporte: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Reporte"

    hoja["A1"] = reporte["titulo"]
    hoja["A1"].font = Font(name="Calibri", size=14, bold=True, color="6E321F")
    hoja["A2"] = f"Generado el {reporte['generado_en']}"
    hoja["A2"].font = Font(name="Calibri", size=9, color="8A7567")

    anchos: dict[int, int] = {}
    fila_actual = 4

    for seccion in reporte["secciones"]:
        celda = hoja.cell(row=fila_actual, column=1, value=seccion["titulo"])
        celda.font = Font(name="Calibri", size=12, bold=True, color="73351F")
        fila_actual += 1

        for columna, encabezado in enumerate(seccion["headers"], start=1):
            celda = hoja.cell(row=fila_actual, column=columna, value=encabezado)
            celda.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", start_color="A14F33", end_color="A14F33")
            celda.alignment = Alignment(horizontal="left")
            anchos[columna] = max(anchos.get(columna, 10), len(str(encabezado)) + 4)
        fila_actual += 1

        for fila in seccion["rows"]:
            for columna, valor in enumerate(fila, start=1):
                celda = hoja.cell(row=fila_actual, column=columna)
                _escribir_celda(celda, valor)
                celda.font = Font(name="Calibri", size=10)
                anchos[columna] = max(anchos.get(columna, 10), min(len(str(valor)) + 3, 50))
            fila_actual += 1

        fila_actual += 2

    for columna, ancho in anchos.items():
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    hoja.freeze_panes = "A5"
    libro.save(buffer)
    buffer.seek(0)
    return buffer

if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(
        debug=debug,
        host=os.environ.get("WEB_HOST", "0.0.0.0"),
        port=int(os.environ.get("WEB_PORT", "5000")),
    )
